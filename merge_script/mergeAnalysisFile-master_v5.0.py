import sys
import os
import getopt
import pandas
import openpyxl
import time
from tqdm import tqdm

start_time = time.time()

def printUsage(lineno):
    print("usage: %s <-r|--ref_file <old file>> <-n|--new_file <new file>> [-m|--mode <Ac|Ar|Setup mode>] [--csv] [--all_case] [--red_zone] (%s:%d)" % (sys.argv[0], os.path.abspath(__file__), lineno))
    print('''   
    -r|--ref_file   :   specify the analyzed excel file
    -n|--new_file   :   specify the current result file witch need be marked
    -m|--mode       :   specify the mark result mode: Ac or Ar or Setup or Syntax, [default: Ac]
    --csv           :   specify the current result file is the csv file, [default: excel file]
    --all_case      :   retain analysis data of the reduced cases, [default: False]
    --red_zone      :   merge the red zone sheets, [default: False]
    -h              :   show help manual
    ''')
    sys.exit(-1)

def defKeyTitles(mode):
    if mode == "Ac":
        # titleListToSearch = ["enno_source", "enno_dest",   "enno_src_clk",  "enno_dest_clk",    "sg_source",   "sg_dest",     "sg_src_clk",    "sg_dest_clk", "enno_file_name",    "sg_file_name",    "enno_reason",   "enno_msgid", "enno_line_num",    "sg_reason",     "sg_rule",    "sg_line_num",      "test_name"]
        if mergeMultiSheets:
            titleListToSearch = ["enno_source", "enno_dest",   "enno_src_clk",  "enno_dest_clk",    "sg_source",   "sg_dest",     "sg_src_clk",    "sg_dest_clk",    "enno_reason",   "enno_msgid", "enno_line_num",    "sg_reason",     "sg_rule",    "sg_line_num",      "test_name"]
        else:
            titleListToSearch = ["enno_source", "enno_dest",   "enno_src_clk",  "enno_dest_clk",    "sg_source",   "sg_dest",     "sg_src_clk",    "sg_dest_clk", "enno_file_name",    "sg_file_name",    "enno_reason",   "enno_msgid",    "sg_reason",     "sg_rule", "test_name"]
    elif mode == "Ar":
        if mergeMultiSheets:
            titleListToSearch = ["Object_list",    "Reason",   "enno_id", "enno_line_num",   "sg_rule",  "sg_line_num", "test_name"]
        else:
            titleListToSearch = ["Object_list",    "enno_file_name",    "sg_file_name",    "Reason",   "enno_id", "enno_line_num",   "sg_rule",  "sg_line_num", "test_name"]
    elif mode == "Setup":
        if mergeMultiSheets:
            titleListToSearch = ["obj_list",    "enno_msgid", "enno_line_num",   "sg_rule",  "sg_line_num", "test_name"]
        else:
            titleListToSearch = ["obj_list",    "enno_file_name",    "sg_file_name",   "enno_msgid", "enno_line_num",   "sg_rule",  "sg_line_num", "test_name"]
    elif mode == "Syntax":
        if mergeMultiSheets:
            titleListToSearch = ['obj_list',   "enno_msgid", "enno_line_num",   "sg_rule",  "sg_line_num", "test_name"]
        else:
            titleListToSearch = ['obj_list',    "enno_file_name",    "sg_file_name",   "enno_msgid", "enno_line_num",   "sg_rule",  "sg_line_num", "test_name"]
    return titleListToSearch

def convert_to_int(variable):
    try:
        return int(variable)
    except ValueError:
        return variable

def get_all_sheets(file_path):
    try:
        all_sheets = pandas.read_excel(file_path, sheet_name=None)
        return all_sheets
    except FileNotFoundError:
        print(f'文件 "{file_path}" 不存在, {os.path.abspath(__file__)}:{sys._getframe().f_lineno}')
        return None
    except Exception as e:
        print(f'读取文件时发生错误： {e}, {os.path.abspath(__file__)}:{sys._getframe().f_lineno}')
        return None

def writeRowFromNew(resultSheet,resultTitlesDict,writeRowIndex,newTitleList,newMessageData):
    # writeRowFromNew()
    for titleName in newTitleList:
        # print (titleName)
        writeCell = resultTitlesDict[titleName] + str(writeRowIndex)
        # resultSheet[writeCell].value = newMessageData[titleName]
        text = newMessageData[titleName]
        if '\1' in str(text):
            resultSheet[writeCell].value = repr(text)
        else:
            resultSheet[writeCell].value = text

def writeRowFromOld(resultSheet,resultTitlesDict,writeRowIndex,newTitleList,oldTitleList,newMessageData,relatedData,redZone):
    # writeRowFromOld()
    toCompareTitleList = ['is_analysis','diff_basis','running_flag']
    if redZone:
        onlyFromOldTitleList = []
    else:
        onlyFromOldTitleList = ['owner','version','SG分析描述','SG执行结果','Enno分析描述','Enno执行结果']
    for resultTitleName in resultTitlesDict.keys():
        writeCell = resultTitlesDict[resultTitleName] + str(writeRowIndex)
        if resultTitleName in toCompareTitleList and resultTitleName in oldTitleList and resultTitleName in newTitleList:
            oldCellValue = relatedData[resultTitleName]
            newCellValue = newMessageData[resultTitleName]
            if pandas.notna(oldCellValue) and pandas.notna(newCellValue):
                if oldCellValue != newCellValue:
                    if resultTitleName == 'is_analysis' and convert_to_int(oldCellValue) == convert_to_int(newCellValue):
                        cellValue = newCellValue
                    else:
                        cellValue = f"{newCellValue}-??-{oldCellValue}"
                else:
                    cellValue = newCellValue
            elif pandas.notna(oldCellValue) and pandas.isna(newCellValue):
                cellValue = oldCellValue
            elif pandas.isna(oldCellValue) and pandas.notna(newCellValue):
                cellValue = newCellValue
            else:
                cellValue = ""
                
            resultSheet[writeCell].value = cellValue
        elif resultTitleName in newTitleList and (resultTitleName not in oldTitleList or resultTitleName not in toCompareTitleList):
            if "\1" in str(newMessageData[resultTitleName]):
                resultSheet[writeCell].value = repr(newMessageData[resultTitleName])
            else:
                resultSheet[writeCell].value = newMessageData[resultTitleName]
        elif resultTitleName in oldTitleList and resultTitleName not in newTitleList:
            cellValue = relatedData[resultTitleName]
            if pandas.notna(cellValue):
                resultSheet[writeCell].value = cellValue
            else:
                resultSheet[writeCell].value = ''

def captureAndWrite(messageLine,oldData,titleListToSearch,resultSheet,resultTitlesDict,writeRowIndex,newTitleList,mode,redZone):
    # find related message row
    relatedData = oldData
    for titleItem in titleListToSearch:
        if pandas.isna(messageLine[titleItem]):
            relatedData = relatedData[relatedData[titleItem].isna()]
        else:
            if '\1' in str(messageLine[titleItem]):
                relatedData = relatedData[
                    (relatedData[titleItem] == repr(messageLine[titleItem]))
                ]
            else:
                relatedData = relatedData[
                    (relatedData[titleItem] == messageLine[titleItem])
                ]
        if relatedData.empty:
            break
    relatedLinesIndexList = relatedData.index
    if len(relatedLinesIndexList) == 0:
        writeRowFromNew(resultSheet,resultTitlesDict,writeRowIndex,newTitleList,messageLine)
    elif len(relatedLinesIndexList) >= 1:
        # writeRowFromOld()
        for i in range(len(relatedLinesIndexList)):
            relatedDataItem = relatedData.iloc[i]
            isRelated = 0
            if relatedDataItem['result'] != mode:
                continue
            else:
                isRelated = 1
                writeRowFromOld(resultSheet,resultTitlesDict,writeRowIndex,newTitleList,oldTitleList,messageLine,relatedDataItem,redZone)
                break
        if isRelated == 0:
            writeRowFromNew(resultSheet,resultTitlesDict,writeRowIndex,newTitleList,messageLine)

def captureCompareAndWrite(messageLine,sheets,titleListToSearch,resultSheet,resultTitlesDict,writeRowIndex,newTitleList,mode,redZone):
    mergeOK = False
    analysisReport = pandas.Series()
    for sheet_name, oldData in sheets.items():
        relatedData = oldData
        for titleItem in titleListToSearch:
            if relatedData.empty:
                break
            if pandas.isna(messageLine[titleItem]):
                relatedData = relatedData[relatedData[titleItem].isna()]
            else:
                relatedData = relatedData[
                    (relatedData[titleItem] == messageLine[titleItem])
                ]
        relatedLinesIndexList = relatedData.index
        if len(relatedLinesIndexList) == 0:
            writeRowFromNew(resultSheet,resultTitlesDict,writeRowIndex,newTitleList,messageLine)
        elif len(relatedLinesIndexList) >= 1:
            # writeRowFromOld()
            for i in range(len(relatedLinesIndexList)):
                relatedDataItem = relatedData.iloc[i]
                isRelated = 0
                if relatedDataItem['result'] != mode:
                    continue
                else:
                    isRelated = 1
                    if relatedDataItem[['SG分析描述', 'SG执行结果', 'Enno分析描述', 'Enno执行结果']].notna().all():
                        if len(analysisReport) == 0:
                            writeRowFromOld(resultSheet,resultTitlesDict,writeRowIndex,newTitleList,oldTitleList,messageLine,relatedDataItem,redZone)
                            mergeOK = True
                        else:
                            if relatedDataItem['SG分析描述'] == analysisReport['SG分析描述'] and relatedDataItem['Enno分析描述'] == analysisReport['Enno分析描述']:
                                analysisReport['SG执行结果'] = relatedDataItem['SG执行结果']
                                analysisReport['Enno执行结果'] = relatedDataItem['Enno执行结果']
                                writeRowFromOld(resultSheet,resultTitlesDict,writeRowIndex,newTitleList,oldTitleList,messageLine,analysisReport,redZone)
                                mergeOK = True
                            elif analysisReport[['SG分析描述', 'SG执行结果', 'Enno分析描述', 'Enno执行结果']].isna().all() and relatedDataItem['is_analysis'] == analysisReport['is_analysis'] and relatedDataItem['diff_basis'] == analysisReport['diff_basis']:
                                analysisReport['owner'] = relatedDataItem['owner']
                                analysisReport['version'] = relatedDataItem['version']
                                analysisReport['SG分析描述'] = relatedDataItem['SG分析描述']
                                analysisReport['SG执行结果'] = relatedDataItem['SG执行结果']
                                analysisReport['Enno分析描述'] = relatedDataItem['Enno分析描述']
                                analysisReport['Enno执行结果'] = relatedDataItem['Enno执行结果']
                                writeRowFromOld(resultSheet,resultTitlesDict,writeRowIndex,newTitleList,oldTitleList,messageLine,analysisReport,redZone)
                                mergeOK = True
                    else:
                        if len(analysisReport) == 0:
                            analysisReport = relatedDataItem
                    break
            if isRelated == 0:
                writeRowFromNew(resultSheet,resultTitlesDict,writeRowIndex,newTitleList,messageLine)
                mergeOK = True
        if mergeOK:
            break
    if not mergeOK and len(analysisReport) != 0:
        writeRowFromOld(resultSheet,resultTitlesDict,writeRowIndex,newTitleList,oldTitleList,messageLine,analysisReport,redZone)


def writeReducedCase(resultSheet,resultTitlesDict,oldData,testCaseName,oldTitleList):
    writeReducedCaseData = oldData[oldData['test_name'] == testCaseName]
    for i in range(len(writeReducedCaseData.index)):
        writeReducedCaseDataItem = writeReducedCaseData.iloc[i]
        writeRowIndex = resultSheet.max_row + 1
        for titleName in oldTitleList:
            writeCell = resultTitlesDict[titleName] + str(writeRowIndex)
            resultSheet[writeCell].value = writeReducedCaseDataItem[titleName]

def writeReducedCases(resultSheet,resultTitlesDict,oldData,redusedCaseList,oldTitleList):
    writeReducedCaseData = oldData[oldData['test_name'].isin(redusedCaseList)]
    for i in tqdm(range(len(writeReducedCaseData.index))):
        writeReducedCaseDataItem = writeReducedCaseData.iloc[i]
        writeRowIndex = resultSheet.max_row + 1
        for titleName in oldTitleList:
            if titleName in resultTitlesDict.keys():
                writeCell = resultTitlesDict[titleName] + str(writeRowIndex)
                resultSheet[writeCell].value = writeReducedCaseDataItem[titleName]

csvType = 0
refFile = ""
newFile = ""
mode = "Ac"
allCase = False
redZone = False
mergeMultiSheets = False

try:
    opts,args =  getopt.getopt(sys.argv[1:],"hr:n:m:",["ref_file=","new_file=","mode=","csv","all_case","red_zone", "merge_multi_sheets"])
except getopt.GetoptError:
    printUsage(sys._getframe().f_lineno)

for opt,arg in opts:
    if opt == '-h':
        printUsage(sys._getframe().f_lineno)
    elif opt in ("-r","--ref_file"):
        refFile = arg
    elif opt in ("-n", "--new_file"):
        newFile = arg
    elif opt in ("-m", "--mode"):
        mode = arg
    elif opt == "--csv":
        csvType = 1
    elif opt == "--all_case":
        allCase = True
    elif opt == "--red_zone":
        redZone = True
    elif opt == "--merge_multi_sheets":
        mergeMultiSheets = True

if refFile == "" or newFile == "":
    printUsage(sys._getframe().f_lineno)

if csvType == 1:
    newData = pandas.read_csv(newFile,encoding='utf-8')
else:
    newData = pandas.read_excel(newFile)

if os.path.splitext(os.path.basename(refFile))[1] == '.xlsx':
    refXlsx = refFile
    if not mergeMultiSheets:
        oldData = pandas.read_excel(refXlsx)
    elif mergeMultiSheets:
        sheets = get_all_sheets(refXlsx)
        if sheets is None:
            print(f'Error: Invalid refrence file "{refXlsx}", {os.path.abspath(__file__)}:{sys._getframe().f_lineno}')
            printUsage(sys._getframe().f_lineno)
        else:
            oldData = sheets[list(sheets.keys())[0]]
elif os.path.splitext(os.path.basename(refFile))[1] == '.csv':
    refXlsx = f'{os.path.splitext(os.path.basename(refFile))[0]}.xlsx'
    oldData = pandas.read_csv(refFile)
    oldData.to_excel(refXlsx, index=False)

newData['enno_line_num'] = [convert_to_int(var) for var in newData['enno_line_num']]
newData['sg_line_num'] = [convert_to_int(var) for var in newData['sg_line_num']]
oldData['enno_line_num'] = [convert_to_int(var) for var in oldData['enno_line_num']]
oldData['sg_line_num'] = [convert_to_int(var) for var in oldData['sg_line_num']]

oldWB = openpyxl.load_workbook(refXlsx)
oldWS = oldWB.active

newTitleList = newData.columns.values
oldTitleList = oldData.columns.values
if redZone:
    toTitles = list(newTitleList)
else:
    toTitles = list(newTitleList)
    for oldTitle in oldTitleList:
        if oldTitle in toTitles:
            continue
        elif oldTitle not in toTitles:
            toTitles.append(oldTitle)

resultSheet = oldWB.create_sheet(os.path.splitext(os.path.basename(newFile))[0],index=0)
resultSheet.append(toTitles)

# 创建字典：titles对应的列序号
resultTitlesDict = {}
for titleCol in resultSheet["1"]:
    resultTitlesDict[titleCol.value] = titleCol.column_letter

# 遍历旧表
titleListToSearch = defKeyTitles(mode)
if not mergeMultiSheets:
    for index in tqdm(newData.index):
        writeRowIndex = index + 2
        messageLine = newData.iloc[index]
        if messageLine['result'] != 'False report' and messageLine['result'] != 'Missing report' and messageLine['result'] != 'Unmatch':
            # writeRowFromNew()
            writeRowFromNew(resultSheet,resultTitlesDict,writeRowIndex,newTitleList,messageLine)
        elif messageLine['result'] == 'False report':
            captureAndWrite(messageLine,oldData,titleListToSearch,resultSheet,resultTitlesDict,writeRowIndex,newTitleList,'False report',redZone)
        elif messageLine['result'] == 'Missing report':
            captureAndWrite(messageLine,oldData,titleListToSearch,resultSheet,resultTitlesDict,writeRowIndex,newTitleList,'Missing report',redZone)
        elif messageLine['result'] == 'Unmatch':
            captureAndWrite(messageLine,oldData,titleListToSearch,resultSheet,resultTitlesDict,writeRowIndex,newTitleList,'Unmatch',redZone)
elif mergeMultiSheets:
    for index in tqdm(newData.index):
        writeRowIndex = index + 2
        messageLine = newData.iloc[index]
        if messageLine['result'] != 'False report' and messageLine['result'] != 'Missing report' and messageLine['result'] != 'Unmatch':
            writeRowFromNew(resultSheet,resultTitlesDict,writeRowIndex,newTitleList,messageLine)
        elif messageLine['result'] == 'False report':
            captureCompareAndWrite(messageLine,sheets,titleListToSearch,resultSheet,resultTitlesDict,writeRowIndex,newTitleList,'False report',redZone)
        elif messageLine['result'] == 'Missing report':
            captureCompareAndWrite(messageLine,sheets,titleListToSearch,resultSheet,resultTitlesDict,writeRowIndex,newTitleList,'Missing report',redZone)
        elif messageLine['result'] == 'Unmatch':
            captureCompareAndWrite(messageLine,sheets,titleListToSearch,resultSheet,resultTitlesDict,writeRowIndex,newTitleList,'Unmatch',redZone)
# oldWB.save(f'{os.path.splitext(os.path.basename(newFile))[0]}.xlsx')

if not mergeMultiSheets:
    if 'test_name' in oldData.columns and 'test_name' in newData.columns:
        oldCaseList = oldData['test_name'].dropna().unique()
        newCaseList = newData['test_name'].dropna().unique()
        addedCaseList = set(newCaseList) - set(oldCaseList)
        redusedCaseList = set(oldCaseList) - set(newCaseList)
        print (f'only exist in {newFile}:{len(addedCaseList)}\n{addedCaseList}')
        print (f'only exist in {refFile}:{len(redusedCaseList)}\n{redusedCaseList}')
        if allCase:
            print (f'start to retain the analysis data of the reduced cases..., {os.path.abspath(__file__)}:{sys._getframe().f_lineno}')
            # if len(redusedCaseList) > 0:
            #     for redusedCase in tqdm(redusedCaseList):
            #         writeReducedCase(resultSheet=resultSheet,resultTitlesDict=resultTitlesDict,oldData=oldData,testCaseName=redusedCase,oldTitleList=oldTitleList)
            writeReducedCases(resultSheet=resultSheet,resultTitlesDict=resultTitlesDict,oldData=oldData,redusedCaseList=redusedCaseList,oldTitleList=oldTitleList)

print (f"Start to save result file..., {os.path.abspath(__file__)}:{sys._getframe().f_lineno}")
if redZone:
    data = resultSheet.values
    columns = next(data)
    df = pandas.DataFrame(data, columns=columns)
    df.to_csv(newFile, index=False, mode='w')
else:
    oldWB.save(f'D:/forBase/mergeMark/result/{os.path.splitext(os.path.basename(newFile))[0]}.xlsx')

end_time = time.time()
print("程序运行时间为：{:.2f}秒".format(end_time - start_time))
