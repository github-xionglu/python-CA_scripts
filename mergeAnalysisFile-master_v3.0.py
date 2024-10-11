import sys
import os
import getopt
import pandas
import openpyxl
import time
from tqdm import tqdm

start_time = time.time()

def printUsage():
    print("usage: %s <-r|--ref_file <old file>> <-n|--new_file <new file>> [-m|--mode <Ac|Ar|Setup mode>] [--csv] [--all_case] [--red_zone]" % (sys.argv[0]))
    print('''   
    -r|--ref_file   :   specify the analyzed excel file
    -n|--new_file   :   specify the current result file witch need be marked
    -m|--mode       :   specify the mark result mode: Ac or Ar or Setup or Syntax, [default: Ac]
    --csv           :   specify the current result file is the csv file, [default: excel file]
    --all_case      :   retain analysis data of the reduced cases, [default: False]
    --red_zone      :   merge the red zone sheets, [default: False]
    -h              :   show help manual
    ''')
    sys.exit(-1)

def defKeyTitles(mode):
    if mode == "Ac":
        # titleListToSearch = ["enno_source", "enno_dest",   "enno_src_clk",  "enno_dest_clk",    "sg_source",   "sg_dest",     "sg_src_clk",    "sg_dest_clk", "enno_file_name",    "sg_file_name",    "enno_reason",   "enno_msgid", "enno_line_num",    "sg_reason",     "sg_rule",    "sg_line_num"]
        titleListToSearch = ["enno_source", "enno_dest",   "enno_src_clk",  "enno_dest_clk",    "sg_source",   "sg_dest",     "sg_src_clk",    "sg_dest_clk", "enno_file_name",    "sg_file_name",    "enno_reason",   "enno_msgid",    "sg_reason",     "sg_rule", "test_name"]
    elif mode == "Ar":
        titleListToSearch = ["Object_list",    "enno_file_name",    "sg_file_name",    "Reason",   "enno_id", "enno_line_num",   "sg_rule",  "sg_line_num", "test_name"]
    elif mode == "Setup":
        titleListToSearch = ["obj_list",    "enno_file_name",    "sg_file_name",   "enno_msgid", "enno_line_num",   "sg_rule",  "sg_line_num", "test_name"]
    elif mode == "Syntax":
        titleListToSearch = ['obj_list',    "enno_file_name",    "sg_file_name",   "enno_msgid", "enno_line_num",   "sg_rule",  "sg_line_num", "test_name"]
    return titleListToSearch

def writeRowFromNew(resultSheet,resultTitlesDict,writeRowIndex,newTitleList,newMessageData):
    # writeRowFromNew()
    for titleName in newTitleList:
        # print (titleName)
        writeCell = resultTitlesDict[titleName] + str(writeRowIndex)
        # resultSheet[writeCell].value = newMessageData[titleName]
        text = newMessageData[titleName]
        if str(text).strip().startswith(''):
            resultSheet[writeCell].value = repr(text)
        else:
            resultSheet[writeCell].value = text

def writeRowFromOld(resultSheet,resultTitlesDict,writeRowIndex,newTitleList,oldTitleList,newMessageData,relatedData,redZone):
    # writeRowFromOld()
    toCompareTitleList = ['is_analysis','diff_basis','running_flag']
    if redZone:
        onlyFromOldTitleList = []
    else:
        onlyFromOldTitleList = ['owner','version','SG分析描述','SG执行结果','Enno分析描述','Enno执行结果']
    for resultTitleName in resultTitlesDict.keys():
        writeCell = resultTitlesDict[resultTitleName] + str(writeRowIndex)
        if resultTitleName in toCompareTitleList and resultTitleName in oldTitleList and resultTitleName in newTitleList:
            oldCellValue = relatedData[resultTitleName]
            newCellValue = newMessageData[resultTitleName]
            if pandas.notna(oldCellValue) and pandas.notna(newCellValue):
                if oldCellValue != newCellValue:
                    cellValue = f"{newCellValue}-??-{oldCellValue}"
                else:
                    cellValue = newCellValue
            elif pandas.notna(oldCellValue) and pandas.isna(newCellValue):
                cellValue = oldCellValue
            elif pandas.isna(oldCellValue) and pandas.notna(newCellValue):
                cellValue = newCellValue
            else:
                cellValue = ""
                
            resultSheet[writeCell].value = cellValue
        elif resultTitleName in newTitleList and (resultTitleName not in oldTitleList or resultTitleName not in toCompareTitleList):
            resultSheet[writeCell].value = newMessageData[resultTitleName]
        elif resultTitleName in oldTitleList and resultTitleName not in newTitleList:
            cellValue = relatedData[resultTitleName]
            if pandas.notna(cellValue):
                resultSheet[writeCell].value = cellValue
            else:
                resultSheet[writeCell].value = ''

def captureAndWrite(messageLine,oldData,titleListToSearch,resultSheet,resultTitlesDict,writeRowIndex,newTitleList,mode,redZone):
    # find related message row
    relatedData = oldData
    for titleItem in titleListToSearch:
        if pandas.isna(messageLine[titleItem]):
            relatedData = relatedData[relatedData[titleItem].isna()]
        else:
            relatedData = relatedData[
                (relatedData[titleItem] == messageLine[titleItem])
            ]
        if relatedData.empty:
            break
    relatedLinesIndexList = relatedData.index
    if len(relatedLinesIndexList) == 0:
        writeRowFromNew(resultSheet,resultTitlesDict,writeRowIndex,newTitleList,messageLine)
    elif len(relatedLinesIndexList) >= 1:
        # writeRowFromOld()
        for i in range(len(relatedLinesIndexList)):
            relatedDataItem = relatedData.iloc[i]
            isRelated = 0
            if relatedDataItem['result'] != mode:
                continue
            else:
                isRelated = 1
                writeRowFromOld(resultSheet,resultTitlesDict,writeRowIndex,newTitleList,oldTitleList,messageLine,relatedDataItem,redZone)
                break
        if isRelated == 0:
            writeRowFromNew(resultSheet,resultTitlesDict,writeRowIndex,newTitleList,messageLine)

def writeReducedCase(resultSheet,resultTitlesDict,oldData,testCaseName,oldTitleList):
    writeReducedCaseData = oldData[oldData['test_name'] == testCaseName]
    for i in range(len(writeReducedCaseData.index)):
        writeReducedCaseDataItem = writeReducedCaseData.iloc[i]
        writeRowIndex = resultSheet.max_row + 1
        for titleName in oldTitleList:
            writeCell = resultTitlesDict[titleName] + str(writeRowIndex)
            resultSheet[writeCell].value = writeReducedCaseDataItem[titleName]

def writeReducedCases(resultSheet,resultTitlesDict,oldData,redusedCaseList,oldTitleList):
    writeReducedCaseData = oldData[oldData['test_name'].isin(redusedCaseList)]
    for i in tqdm(range(len(writeReducedCaseData.index))):
        writeReducedCaseDataItem = writeReducedCaseData.iloc[i]
        writeRowIndex = resultSheet.max_row + 1
        for titleName in oldTitleList:
            if titleName in resultTitlesDict.keys():
                writeCell = resultTitlesDict[titleName] + str(writeRowIndex)
                resultSheet[writeCell].value = writeReducedCaseDataItem[titleName]

csvType = 0
refXlsx = ""
newFile = ""
mode = "Ac"
allCase = False
redZone = False

try:
    opts,args =  getopt.getopt(sys.argv[1:],"hr:n:m:",["ref_file=","new_file=","mode=","csv","all_case","red_zone"])
except getopt.GetoptError:
    printUsage()

for opt,arg in opts:
    if opt == '-h':
        printUsage()
    elif opt in ("-r","--ref_file"):
        refXlsx = arg
    elif opt in ("-n", "--new_file"):
        newFile = arg
    elif opt in ("-m", "--mode"):
        mode = arg
    elif opt == "--csv":
        csvType = 1
    elif opt == "--all_case":
        allCase = True
    elif opt == "--red_zone":
        redZone = True

if refXlsx == "" or newFile == "":
    printUsage()

if csvType == 1:
    newData = pandas.read_csv(newFile,encoding='utf-8')
else:
    newData = pandas.read_excel(newFile)
oldData = pandas.read_excel(refXlsx)

oldWB = openpyxl.load_workbook(refXlsx)
oldWS = oldWB.active

newTitleList = newData.columns.values
oldTitleList = oldData.columns.values
if redZone:
    toTitles = list(newTitleList)
else:
    toTitles = list(newTitleList)
    for oldTitle in oldTitleList:
        if oldTitle in toTitles:
            continue
        elif oldTitle not in toTitles:
            toTitles.append(oldTitle)

resultSheet = oldWB.create_sheet(os.path.splitext(os.path.basename(newFile))[0],index=0)
resultSheet.append(toTitles)

# 创建字典：titles对应的列序号
resultTitlesDict = {}
for titleCol in resultSheet["1"]:
    resultTitlesDict[titleCol.value] = titleCol.column_letter

# 遍历旧表
titleListToSearch = defKeyTitles(mode)

for index in tqdm(newData.index):
    writeRowIndex = index + 2
    messageLine = newData.iloc[index]
    if messageLine['result'] != 'False report' and messageLine['result'] != 'Missing report' and messageLine['result'] != 'Unmatch':
        # writeRowFromNew()
        writeRowFromNew(resultSheet,resultTitlesDict,writeRowIndex,newTitleList,messageLine)
    elif messageLine['result'] == 'False report':
        captureAndWrite(messageLine,oldData,titleListToSearch,resultSheet,resultTitlesDict,writeRowIndex,newTitleList,'False report',redZone)
    elif messageLine['result'] == 'Missing report':
        captureAndWrite(messageLine,oldData,titleListToSearch,resultSheet,resultTitlesDict,writeRowIndex,newTitleList,'Missing report',redZone)
    elif messageLine['result'] == 'Unmatch':
        captureAndWrite(messageLine,oldData,titleListToSearch,resultSheet,resultTitlesDict,writeRowIndex,newTitleList,'Unmatch',redZone)
    
# oldWB.save(f'{os.path.splitext(os.path.basename(newFile))[0]}.xlsx')

if 'test_name' in oldData.columns and 'test_name' in newData.columns:
    oldCaseList = oldData['test_name'].dropna().unique()
    newCaseList = newData['test_name'].dropna().unique()
    addedCaseList = set(newCaseList) - set(oldCaseList)
    redusedCaseList = set(oldCaseList) - set(newCaseList)
    print (f'only exist in {newFile}:{len(addedCaseList)}\n{addedCaseList}')
    print (f'only exist in {refXlsx}:{len(redusedCaseList)}\n{redusedCaseList}')
    if allCase:
        print (f'start to retain the analysis data of the reduced cases...')
        # if len(redusedCaseList) > 0:
        #     for redusedCase in tqdm(redusedCaseList):
        #         writeReducedCase(resultSheet=resultSheet,resultTitlesDict=resultTitlesDict,oldData=oldData,testCaseName=redusedCase,oldTitleList=oldTitleList)
        writeReducedCases(resultSheet=resultSheet,resultTitlesDict=resultTitlesDict,oldData=oldData,redusedCaseList=redusedCaseList,oldTitleList=oldTitleList)

oldWB.save(f'D:/forBase/mergeMark/result/{os.path.splitext(os.path.basename(newFile))[0]}.xlsx')

end_time = time.time()
print("程序运行时间为：{:.2f}秒".format(end_time - start_time))
