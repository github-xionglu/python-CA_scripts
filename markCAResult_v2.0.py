import sys
import openpyxl
import time
import os
import shutil
import re
import pandas
import getopt
import csv
import codecs
from tqdm import tqdm
import inspect
import subprocess

def printUsage(lineno):
    print("usage: %s <-r|--ref <ref file>> [-e|--enno <enno result file>] [-s|--sg <sg result file>] [-m|--mode <Ac|Ar|Setup mode>] [--csv] [--appendLabel] [--inferStatus <any|infer|uninfer>] [--dealDoubleQuotation] (%s:%d)" % (sys.argv[0], os.path.abspath(__file__), lineno))
    print('''   
    -r|--ref                :   specify the reference file
    -e|--enno               :   specify the enno result file witch need be marked
    -s|--sg                 :   specify the sg result file witch need be marked
    -m|--mode               :   specify the mark result mode: Ac or Ar or Setup. [default: Ac]
    --csv                   :   specify the input files are the csv file
    --appendLabel           :   specify the running mode: mark or appendLabel. [default: mark]
    --inferStatus           :   specify the infer status: any, infer or uninfer. [default: any]
    --dealDoubleQuotation   :   specify the input files are the csv file
    -h                      :   show help manual
    ''')
    sys.exit(-1)

start_time = time.time()


csvType = 0
refFile = ""
ennoFile = ""
sgFile = ""
ennoXlsx = ""
sgXlsx = ""
mode = "Ac"
appendLabel = 0
inferStatus = "any"
dealDoubleQuotation = False

try:
    opts,args =  getopt.getopt(sys.argv[1:],"hr:e:s:m:",["ref=","enno=","sg=","mode=","csv","appendLabel","inferStatus=","dealDoubleQuotation"])
except getopt.GetoptError:
    printUsage(0)

for opt,arg in opts:
    if opt == '-h':
        printUsage(0)
    elif opt in ("-r","--ref"):
        refFile = arg
    elif opt in ("-e", "--enno"):
        ennoFile = arg
    elif opt in ("-s", "--sg"):
        sgFile = arg
    elif opt in ("-m", "--mode"):
        mode = arg
    elif opt == "--csv":
        csvType = 1
    elif opt == "--appendLabel":
        appendLabel = 1
    elif opt == "--inferStatus":
        inferStatus = arg
    elif opt == "--dealDoubleQuotation":
        dealDoubleQuotation = True

if bool(refFile):
    if os.path.exists(refFile):
        refFileName = os.path.basename(refFile)
        if os.path.splitext(refFileName)[1] == ".csv":
            if dealDoubleQuotation:
                optmizeFile = f'{os.path.splitext(refFileName)[0]}_opt.csv'
                shutil.copy(refFile,optmizeFile)
                subprocess.call(["sed", "-i", "s/\"/\"\"/g", optmizeFile])
                subprocess.call(["sed", "-i", "s/,\"\"/,\"/g; s/\"\",/\",/g; s/\"\"$/\"/g", optmizeFile])
                refData = pandas.read_csv(optmizeFile,encoding='utf-8',encoding_errors='ignore',sep=',',dtype={'e_message':str,'enno_line_num':int,'sg_line_num':int})
            else:
                refData = pandas.read_csv(refFile,encoding='utf-8',encoding_errors='ignore',sep=',',dtype={'e_message':str,'enno_line_num':int,'sg_line_num':int})
        elif os.path.splitext(refFileName)[1] == ".xlsx":
            refData = pandas.read_excel(refFile)
        else:
            print("Error: invalid reference file type: " + refFile)
            sys.exit(0)
    else:
        print("Error: there is not such file: " + refFile)
        sys.exit(0)
else:
    print("Error: there is reference file, please check it")
    sys.exit(0)

if inferStatus == "any":
    refData = refData
elif inferStatus == "infer":
    refData = refData[~refData['test_name'].str.startswith('uninfer_')]
elif inferStatus == "uninfer":
    refData = refData[refData['test_name'].str.startswith('uninfer_')]
else:
    print(f'Invalid argument {inferStatus} of the option "--inferStatus"')

if not bool(ennoFile) and not bool(sgFile):
    print ("Error: there is not correct input file, enno result file or/and sg result file")
    sys.exit(0)

runOK = 0
if bool(ennoFile):
    if not os.path.exists(ennoFile):
        print("Warn: the " + ennoFile + " is not a file")
    else:
        if csvType == 0:
            ennoXlsx = ennoFile
            ennoXlsxName = os.path.basename(ennoXlsx)
        elif csvType == 1:
            ennoXlsxName = os.path.splitext(ennoFile)[0] + ".xlsx"
            with open(ennoFile,'r') as f:
                reader = csv.reader(f)
                data = list(reader)
            wb = openpyxl.Workbook()
            ws = wb.active
            for row in data:
                # ws.append(row)
                for index in range(len(row)):
                    if '\\x01' in repr(row[index]):
                        row[index] = repr(row[index])
                ws.append(row)
            wb.save(ennoXlsxName)
            ennoXlsx = os.path.abspath(ennoXlsxName)
        runOK = 1
if bool(sgFile):
    if not os.path.exists(sgFile):
        print("Warn: the " + sgFile + " is not a file")
    else:
        if csvType == 0:
            sgXlsx = sgFile
            sgXlsxName = os.path.basename(sgXlsx)
        elif csvType == 1:
            sgXlsxName = os.path.splitext(sgFile)[0] + ".xlsx"
            with open(sgFile,'r') as f:
                reader = csv.reader(f)
                data = list(reader)
            wb = openpyxl.Workbook()
            ws = wb.active
            for row in data:
                for index in range(len(row)):
                    if '\\x01' in repr(row[index]):
                        row[index] = repr(row[index])
                ws.append(row)
            wb.save(sgXlsxName)
            sgXlsx = os.path.abspath(sgXlsxName)
        runOK = 1
if runOK == 0:
    sys.exit(0)




refTitles = refData.columns
if 'diff_basis' not in refTitles or 'file_name' not in refTitles or 'e_message' not in refTitles or 's_message' not in refTitles:
    print("Error: the file " + refFile + " is not a standard workbook")
    print ('refTitles:' + refTitles)
    sys.exit(0)

fromTitle = "from"
fromTag = refFileName

logFile = "mark_ca_result.log"
writeLog = open(logFile,"w")

if bool(ennoXlsx):
    enno_start_time = time.time()
    print (f'ennoXlsx: {ennoXlsx}')
    ennoWB = openpyxl.load_workbook(ennoXlsx,read_only=False)
    ennoWS = ennoWB.active

    ennoTitlesDict = {}
    for titleCol in ennoWS['1']:
        ennoTitlesDict[titleCol.value] = titleCol.column_letter
    if fromTitle not in ennoTitlesDict:
        # 增加from列
        ennoMaxCol = ennoWS.max_column
        ennoWS.insert_cols(ennoMaxCol +  1)
        ennoWS.cell(row=1,column=ennoMaxCol+1,value=fromTitle)

        ennoTitlesDict = {}
        for titleCol in ennoWS["1"]:
            ennoTitlesDict[titleCol.value] = titleCol.column_letter

    for i in tqdm(range(2, ennoWS.max_row + 1)):
        isPass = 1

        ennoResultCell = ennoTitlesDict['Result'] + str(i)
        ennoResultCell_value = ennoWS[ennoResultCell].value
        if appendLabel == 0:
            flag = ennoResultCell_value == "ok" or ennoResultCell_value == "enno_reason_more" or ennoResultCell_value == "undo"
        elif appendLabel == 1:
            flag = not ennoResultCell_value is None
        if flag:
            continue
        # elif ennoResultCell_value is None or ennoResultCell_value != "ok" or ennoResultCell_value != "enno_reason_more" or ennoResultCell_value != "undo":
        else:
            if mode == "Ac":
                ennoMsgIdCell = ennoTitlesDict['Enno_id'] + str(i)
                ennoMsgIdCell_value = ennoWS[ennoMsgIdCell].value
                ennoReasonCell = ennoTitlesDict['ERule'] + str(i)
                ennoReasonCell_value = ennoWS[ennoReasonCell].value
                ennoMsgCell = ennoTitlesDict['Message'] + str(i)
                ennoMsgCell_value = ennoWS[ennoMsgCell].value
                ennoFileCell = ennoTitlesDict['File'] + str(i)
                ennoFileCell_value = ennoWS[ennoFileCell].value
                # ennoLineCell = ennoTitlesDict['Line'] + str(i)
                # ennoLineCell_value = ennoWS[ennoLineCell].value
            elif mode == "Ar":
                ennoMsgIdCell = ennoTitlesDict['Enno_id'] + str(i)
                ennoMsgIdCell_value = ennoWS[ennoMsgIdCell].value
                # ennoReasonCell = ennoTitlesDict['ERule'] + str(i)
                # ennoReasonCell_value = ennoWS[ennoReasonCell].value
                ennoMsgCell = ennoTitlesDict['Message'] + str(i)
                ennoMsgCell_value = ennoWS[ennoMsgCell].value
                ennoFileCell = ennoTitlesDict['File'] + str(i)
                ennoFileCell_value = ennoWS[ennoFileCell].value
                # ennoLineCell = ennoTitlesDict['Line'] + str(i)
                # ennoLineCell_value = ennoWS[ennoLineCell].value
            elif mode == "Setup":
                # ennoMsgIdCell = ennoTitlesDict['Enno_id'] + str(i)
                # ennoMsgIdCell_value = ennoWS[ennoMsgIdCell].value
                ennoReasonCell = ennoTitlesDict['ERule'] + str(i)
                ennoReasonCell_value = ennoWS[ennoReasonCell].value
                ennoMsgCell = ennoTitlesDict['Message'] + str(i)
                ennoMsgCell_value = ennoWS[ennoMsgCell].value
                ennoFileCell = ennoTitlesDict['File'] + str(i)
                ennoFileCell_value = ennoWS[ennoFileCell].value
                # ennoLineCell = ennoTitlesDict['Line'] + str(i)
                # ennoLineCell_value = ennoWS[ennoLineCell].value
            elif mode == "Syntax":
                # ennoMsgIdCell = ennoTitlesDict['Enno_id'] + str(i)
                # ennoMsgIdCell_value = ennoWS[ennoMsgIdCell].value
                ennoMsgIdCell = ennoTitlesDict['ERule'] + str(i)
                ennoMsgIdCell_value = ennoWS[ennoMsgIdCell].value
                # ennoReasonCell = ennoTitlesDict['ERule'] + str(i)
                # ennoReasonCell_value = ennoWS[ennoReasonCell].value
                ennoMsgCell = ennoTitlesDict['Message'] + str(i)
                ennoMsgCell_value = ennoWS[ennoMsgCell].value
                ennoFileCell = ennoTitlesDict['File'] + str(i)
                ennoFileCell_value = ennoWS[ennoFileCell].value
                # ennoLineCell = ennoTitlesDict['Line'] + str(i)
                # ennoLineCell_value = ennoWS[ennoLineCell].value
                ennoSeverityCell = ennoTitlesDict['Severity'] + str(i)
                ennoSeverityCell_value = ennoWS[ennoSeverityCell].value
            else:
                # print ("Warn: invalid mode: " + mode + ", use default mode: Ac")
                ennoMsgIdCell = ennoTitlesDict['Enno_id'] + str(i)
                ennoMsgIdCell_value = ennoWS[ennoMsgIdCell].value
                ennoReasonCell = ennoTitlesDict['ERule'] + str(i)
                ennoReasonCell_value = ennoWS[ennoReasonCell].value
                ennoMsgCell = ennoTitlesDict['Message'] + str(i)
                ennoMsgCell_value = ennoWS[ennoMsgCell].value
                ennoFileCell = ennoTitlesDict['File'] + str(i)
                ennoFileCell_value = ennoWS[ennoFileCell].value
                # ennoLineCell = ennoTitlesDict['Line'] + str(i)
                # ennoLineCell_value = ennoWS[ennoLineCell].value


            if ennoMsgCell_value is None:
                continue

            if mode == "Ac":
                # matchRefObjList = ['file_name','e_message','enno_msgid','enno_reason']
                refMatchLines = refData.loc[
                    (refData['file_name'] == ennoFileCell_value)
                    &
                    (refData['e_message'].isin([ennoMsgCell_value, f'{ennoMsgCell_value} ']))
                    &
                    (refData['enno_msgid'] == ennoMsgIdCell_value)
                    &
                    (refData['enno_reason'] == ennoReasonCell_value)
                ]
            elif mode == "Ar":
                # matchRefObjList = ['file_name','e_message','enno_msgid']
                refMatchLines = refData.loc[
                    (refData['file_name'] == ennoFileCell_value)
                    &
                    (refData['e_message'].isin([ennoMsgCell_value, f'{ennoMsgCell_value} ']))
                    &
                    (refData['enno_id'] == ennoMsgIdCell_value)
                ]
            elif mode == "Setup":
                # matchRefObjList = ['file_name','e_message','enno_msgid','enno_reason']
                # refMatchLines = refData.loc[
                #     (refData['file_name'] == ennoFileCell_value)
                #     &
                #     (refData['e_message'] == ennoMsgCell_value)
                #     &
                #     (refData['enno_msgid'] == ennoMsgIdCell_value)
                #     &
                #     (refData['enno_reason'] == ennoReasonCell_value)
                # ]
                refMatchLines = refData.loc[
                    (refData['file_name'] == ennoFileCell_value)
                    &
                    (refData['e_message'].isin([ennoMsgCell_value, f'{ennoMsgCell_value} ']))
                    &
                    (refData['enno_msgid'] == ennoReasonCell_value)
                ]
            elif mode == "Syntax":
                # matchRefObjList = ['file_name', 'e_message','enno_msgid', 'e_severity']
                refMatchLines = refData.loc[
                    (refData['file_name'] == ennoFileCell_value)
                    &
                    (refData['e_message'].isin([ennoMsgCell_value, f'{ennoMsgCell_value} ']))
                    &
                    (refData['enno_msgid'] == ennoMsgIdCell_value)
                    &
                    (refData['e_severity'] == ennoSeverityCell_value)
                ]
            else:
                # matchRefObjList = ['file_name','e_message','enno_msgid','enno_reason']
                refMatchLines = refData.loc[
                    (refData['file_name'] == ennoFileCell_value)
                    &
                    (refData['e_message'].isin([ennoMsgCell_value, f'{ennoMsgCell_value} ']))
                    &
                    (refData['enno_msgid'] == ennoMsgIdCell_value)
                    &
                    (refData['enno_reason'] == ennoReasonCell_value)
                ]
            refMatchLinesIndexList = refMatchLines.index
            if len(refMatchLinesIndexList) == 0:
                # print("cannot found the message: " + ennoMsgCell_value + ", row num: " + str(i))
                writeLog.write("=====================\ncannot found the message: \n" + ennoMsgCell_value + "\nrow num: " + str(i) + "\n")
                writeLog.write("refMatchLines: \n" + str(refMatchLines) + "\n")
                ennoWS[ennoResultCell].value = ""
                continue
            elif len(refMatchLinesIndexList) >= 1:
                writeLog.write(f'refMatchLines: \n{str(refMatchLines)}\n')
                forCAResult = []
                for refMatchLinesIndex in refMatchLinesIndexList:
                    refResultValue = refData.at[refMatchLinesIndex,'result']
                    if pandas.notna(refResultValue):
                        refResultValue = str(refResultValue).strip()
                    else:
                        isPass = 0
                        break
                    if refResultValue != "Missing report" and refResultValue != "False report" and refResultValue != "Unmatch" and refResultValue != "pass":
                        isPass = isPass
                        forCAResult.append(refResultValue)
                        continue
                    elif refResultValue == "Missing report" or refResultValue == "False report" or refResultValue == "Unmatch" or refResultValue == "pass":
                        refDiffIdValue = refData.at[refMatchLinesIndex,'diff_basis']
                        if pandas.notna(refDiffIdValue):
                            # refDiffIdValue = str(';'.join(str(refDiffIdValue).strip().split()))
                            isPass = isPass
                            forCAResult.extend(refDiffIdValue.split(";"))
                            continue
                        else:
                            refDiffIdValue = refData.at[refMatchLinesIndex,'running_flag']
                            if pandas.notna(refDiffIdValue):
                                #
                                isPass = isPass
                                forCAResult.extend(refDiffIdValue.split(";"))
                            else:
                                isPass = 0
                                break
                if not bool(isPass):
                    ennoWS[ennoResultCell].value = ""
                    continue
                else:
                    ennoFromCell = ennoTitlesDict[fromTitle] + str(i)
                    ennoWS[ennoFromCell].value = fromTag
                    ennoWS[ennoResultCell].value = str(";".join(sorted(set(sorted(forCAResult)))))

    ennoWB.save(ennoXlsx)
    ennoWB.close()

    if csvType == 1:
        ennoFileBak = os.path.splitext(ennoFile)[0] + "_bak.csv"
        os.rename(ennoFile,ennoFileBak)
        df = pandas.read_excel(ennoXlsx)
        df.to_csv(ennoFile,index=False,encoding='utf-8')

    enno_end_time = time.time()
    print("enno反标运行时间为：{:.2f}秒".format(enno_end_time - enno_start_time))




if bool(sgXlsx):
    print(f'sgXlsx: {sgXlsx}')
    sg_start_time = time.time()
    sgWB = openpyxl.load_workbook(sgXlsx,read_only=False)
    sgWS = sgWB.active

    sgTitlesDict = {}
    for titleCol in sgWS["1"]:
        sgTitlesDict[titleCol.value] = titleCol.column_letter
    if fromTitle not in sgTitlesDict:
        # 增加from列
        sgMaxCol = sgWS.max_column
        sgWS.insert_cols(sgMaxCol +  1)
        sgWS.cell(row=1,column=sgMaxCol+1,value=fromTitle)

        sgTitlesDict = {}
        for titleCol in sgWS["1"]:
            sgTitlesDict[titleCol.value] = titleCol.column_letter

    for j in tqdm(range(2, sgWS.max_row + 1)):
        isPass = 1

        sgResultCell = sgTitlesDict['Result'] + str(j)
        sgResultCell_value = sgWS[sgResultCell].value
        if appendLabel == 0:
            flag = sgResultCell_value == "ok" or sgResultCell_value == "enno_reason_more" or sgResultCell_value == "undo"
        elif appendLabel == 1:
            flag = not sgResultCell_value is None
        if flag:
            continue
        else:
            if mode == "Ac":
                # sgMsgIdCell = sgTitlesDict['ARule'] + str(j)
                # sgMsgIdCell_value = sgWS[sgMsgIdCell].value
                sgReasonCell = sgTitlesDict['ARule'] + str(j)
                sgReasonCell_value = sgWS[sgReasonCell].value
                sgMsgCell = sgTitlesDict['Message'] + str(j)
                sgMsgCell_value = sgWS[sgMsgCell].value
                sgFileCell = sgTitlesDict['File'] + str(j)
                sgFileCell_value = sgWS[sgFileCell].value
                # sgLineCell = sgTitlesDict['Line'] + str(j)
                # sgLineCell_value = sgWS[sgLineCell].value
            elif mode == "Ar":
                # sgMsgIdCell = sgTitlesDict['ARule'] + str(j)
                # sgMsgIdCell_value = sgWS[sgMsgIdCell].value
                # sgReasonCell = sgTitlesDict['ARule'] + str(j)
                # sgReasonCell_value = sgWS[sgReasonCell].value
                sgMsgCell = sgTitlesDict['Message'] + str(j)
                sgMsgCell_value = sgWS[sgMsgCell].value
                sgFileCell = sgTitlesDict['File'] + str(j)
                sgFileCell_value = sgWS[sgFileCell].value
                # sgLineCell = sgTitlesDict['Line'] + str(j)
                # sgLineCell_value = sgWS[sgLineCell].value
            elif mode == "Setup":
                # sgMsgIdCell = sgTitlesDict['ARule'] + str(j)
                # sgMsgIdCell_value = sgWS[sgMsgIdCell].value
                sgReasonCell = sgTitlesDict['ARule'] + str(j)
                sgReasonCell_value = sgWS[sgReasonCell].value
                sgMsgCell = sgTitlesDict['Message'] + str(j)
                sgMsgCell_value = sgWS[sgMsgCell].value
                sgFileCell = sgTitlesDict['File'] + str(j)
                sgFileCell_value = sgWS[sgFileCell].value
                sgLineCell = sgTitlesDict['Line'] + str(j)
                sgLineCell_value = sgWS[sgLineCell].value
            elif mode == "Syntax":
                sgMsgIdCell = sgTitlesDict['ARule'] + str(j)
                sgMsgIdCell_value = sgWS[sgMsgIdCell].value
                # sgReasonCell = sgTitlesDict['ARule'] + str(j)
                # sgReasonCell_value = sgWS[sgReasonCell].value
                sgMsgCell = sgTitlesDict['Message'] + str(j)
                sgMsgCell_value = sgWS[sgMsgCell].value
                sgFileCell = sgTitlesDict['File'] + str(j)
                sgFileCell_value = sgWS[sgFileCell].value
                sgLineCell = sgTitlesDict['Line'] + str(j)
                sgLineCell_value = sgWS[sgLineCell].value
                sgSeverityCell = sgTitlesDict['Severity'] + str(j)
                sgSeverityCell_value = sgWS[sgSeverityCell].value
            else:
                # sgMsgIdCell = sgTitlesDict['ARule'] + str(j)
                # sgMsgIdCell_value = sgWS[sgMsgIdCell].value
                sgReasonCell = sgTitlesDict['ARule'] + str(j)
                sgReasonCell_value = sgWS[sgReasonCell].value
                sgMsgCell = sgTitlesDict['Message'] + str(j)
                sgMsgCell_value = sgWS[sgMsgCell].value
                sgFileCell = sgTitlesDict['File'] + str(j)
                sgFileCell_value = sgWS[sgFileCell].value
                # sgLineCell = sgTitlesDict['Line'] + str(j)
                # sgLineCell_value = sgWS[sgLineCell].value

            if sgMsgCell_value is None:
                continue

            if mode == "Ac":
                # matchRefObjList = ['file_name','s_message','sg_reason']
                refMatchLines = refData.loc[
                    (refData['file_name'] == sgFileCell_value)
                    &
                    (refData['s_message'].isin([sgMsgCell_value, f'{sgMsgCell_value} ']))
                    &
                    (refData['sg_reason'] == sgReasonCell_value)
                ]
            elif mode == "Ar":
                # matchRefObjList = ['file_name','s_message']
                refMatchLines = refData.loc[
                    (refData['file_name'] == sgFileCell_value)
                    &
                    (refData['s_message'].isin([sgMsgCell_value, f'{sgMsgCell_value} ']))
                ]
            elif mode == "Setup":
                # matchRefObjList = ['file_name','s_message','sg_reason']
                refMatchLines = refData.loc[
                    (refData['file_name'] == sgFileCell_value)
                    &
                    (refData['s_message'].isin([sgMsgCell_value, f'{sgMsgCell_value} ']))
                    &
                    (refData['sg_rule'] == sgReasonCell_value)
                    &
                    (refData['sg_line_num'] == float(sgLineCell_value))
                ]
            elif mode == "Syntax":
                # matchRefObjList = ['file_name','s_message','sg_rule', 'Severity']
                # refMatchLines = refData.loc[
                #    (refData['file_name'] == sgFileCell_value)
                #    &
                #    (refData['s_message'] == sgMsgCell_value)
                #    &
                #    (refData['sg_rule'] == sgReasonCell_value)
                #    &
                #    (refData['s_severity'].upper() == sgSeverityCell_value)
                # ]
                refData['file_name'] = refData['file_name'].str.replace('.sdc','.sgdc', regex=False)
                refMatchLines = refData.loc[
                    (refData['file_name'] == sgFileCell_value)
                    &
                    (refData['s_message'].isin([sgMsgCell_value, f'{sgMsgCell_value} ']))
                    &
                    (refData['sg_rule'] == sgMsgIdCell_value)
                ]
            else:
                # matchRefObjList = ['file_name','s_message','sg_reason']
                refMatchLines = refData.loc[
                    (refData['file_name'] == sgFileCell_value)
                    &
                    (refData['s_message'].isin([sgMsgCell_value, f'{sgMsgCell_value} ']))
                    &
                    (refData['sg_reason'] == sgReasonCell_value)
                ]
            refMatchLinesIndexList = refMatchLines.index
            if len(refMatchLinesIndexList) == 0:
                writeLog.write("=====================\ncannot found the message: \n" + sgMsgCell_value + "\nrow num: " + str(j) + "\n")
                sgWS[sgResultCell].value = ''
                continue
            elif len(refMatchLinesIndexList) >= 1:
                forCAResult = []
                for refMatchLinesIndex in refMatchLinesIndexList:
                    refResultValue = refData.at[refMatchLinesIndex,'result']
                    if pandas.notna(refResultValue):
                        refResultValue = str(refResultValue).strip()
                    else:
                        isPass = 0
                        break
                    if refResultValue != "Missing report" and refResultValue != "False report" and refResultValue != "Unmatch" and refResultValue != "pass":
                        isPass = isPass
                        forCAResult.append(refResultValue)
                        continue
                    elif refResultValue == "Missing report" or refResultValue == "False report" or refResultValue == "Unmatch" or refResultValue == "pass":
                        refDiffIdValue = refData.at[refMatchLinesIndex,'diff_basis']
                        if pandas.notna(refDiffIdValue):
                            # refDiffIdValue = str(';'.join(str(refDiffIdValue).strip().split()))
                            isPass = isPass
                            forCAResult.extend(refDiffIdValue.split(";"))
                        else:
                            refDiffIdValue = refData.at[refMatchLinesIndex,'running_flag']
                            if pandas.notna(refDiffIdValue):
                                # refDiffIdValue = str(';'.join(str(refDiffIdValue).strip().split()))
                                isPass = isPass
                                forCAResult.extend(refDiffIdValue.split(";"))
                            else:
                                isPass = 0
                                break
                if isPass == 0:
                    sgWS[sgResultCell].value = ""
                    continue
                else:
                    sgFromCell = sgTitlesDict[fromTitle] + str(j)
                    sgWS[sgFromCell].value = fromTag
                    sgWS[sgResultCell].value = str(";".join(sorted(set(sorted(forCAResult)))))

    sgWB.save(sgXlsx)
    sgWB.close()

    if csvType == 1:
        sgFileBak = os.path.splitext(sgFile)[0] + "_bak.csv"
        os.rename(sgFile,sgFileBak)
        df = pandas.read_excel(sgXlsx)
        df.to_csv(sgFile,index=False,encoding='utf-8')

    sg_end_time = time.time()
    print("SG反标运行时间为：{:.2f}秒".format(sg_end_time - sg_start_time))

writeLog.close()
end_time = time.time()
print("程序运行时间为：{:.2f}秒".format(end_time - start_time))
